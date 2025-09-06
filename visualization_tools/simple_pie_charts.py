#!/usr/bin/env python3
"""
Simple script to create pie charts from XLSX files using minimal dependencies.
This script uses openpyxl directly to read Excel files and matplotlib for visualization.
"""

import os
import sys
from pathlib import Path
import argparse
import json
from collections import Counter, defaultdict

# Try to import required packages with fallback options
try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not available. Please install: pip install openpyxl")
    openpyxl = None

try:
    import matplotlib
    matplotlib.use('Agg')  # Use non-interactive backend
    import matplotlib.pyplot as plt
    matplotlib_available = True
except ImportError:
    print("matplotlib not available. Will generate data summaries only.")
    matplotlib_available = False


class SimplePieChartGenerator:
    """Simple pie chart generator using minimal dependencies."""
    
    def __init__(self, output_dir="charts"):
        """Initialize the generator."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def read_xlsx_with_openpyxl(self, file_path):
        """Read XLSX file using openpyxl and return data as list of dictionaries."""
        if not openpyxl:
            print("openpyxl not available, cannot read XLSX files")
            return None
            
        try:
            workbook = load_workbook(file_path, read_only=True)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(str(cell.value) if cell.value is not None else f"Column_{len(headers)}")
            
            # Read data rows
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_data = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            row_data[headers[i]] = cell
                    data.append(row_data)
            
            workbook.close()
            return {'headers': headers, 'data': data}
            
        except Exception as e:
            print(f"Error reading {file_path} with openpyxl: {e}")
            return None
    
    def analyze_column_data(self, data, column):
        """Analyze a column to determine if it's suitable for pie chart."""
        values = []
        for row in data:
            val = row.get(column)
            if val is not None and str(val).strip():
                values.append(str(val))
        
        if not values:
            return None
            
        # Count occurrences
        counter = Counter(values)
        
        # Filter out values that appear very infrequently (less than 1% of total)
        total_count = len(values)
        min_count = max(1, total_count * 0.01)
        
        filtered_counter = {k: v for k, v in counter.items() if v >= min_count}
        
        # Only create chart if we have 2-20 unique values
        if 2 <= len(filtered_counter) <= 20:
            return filtered_counter
        
        return None
    
    def create_pie_chart(self, data_dict, title, filename):
        """Create a pie chart from data dictionary."""
        if not matplotlib_available:
            print(f"Matplotlib not available. Data summary for {title}:")
            total = sum(data_dict.values())
            for label, count in sorted(data_dict.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total) * 100
                print(f"  {label}: {count} ({percentage:.1f}%)")
            print()
            return False
        
        try:
            plt.figure(figsize=(10, 8))
            
            labels = list(data_dict.keys())
            values = list(data_dict.values())
            
            # Create pie chart
            wedges, texts, autotexts = plt.pie(
                values,
                labels=labels,
                autopct='%1.1f%%',
                startangle=90,
                textprops={'fontsize': 9}
            )
            
            # Improve text readability
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
            
            plt.title(title, fontsize=14, fontweight='bold')
            plt.axis('equal')
            
            # Save the chart
            plt.savefig(filename, dpi=300, bbox_inches='tight')
            plt.close()
            
            return True
            
        except Exception as e:
            print(f"Error creating chart {filename}: {e}")
            return False
    
    def process_numeric_by_category(self, data, numeric_col, category_col):
        """Process numeric data grouped by category."""
        grouped_data = defaultdict(float)
        
        for row in data:
            category = row.get(category_col)
            numeric_val = row.get(numeric_col)
            
            if category is not None and numeric_val is not None:
                try:
                    # Try to convert to float
                    if isinstance(numeric_val, (int, float)):
                        num_val = float(numeric_val)
                    else:
                        # Try to parse string as number
                        cleaned = str(numeric_val).replace(',', '').replace('$', '').strip()
                        num_val = float(cleaned)
                    
                    grouped_data[str(category)] += num_val
                    
                except (ValueError, TypeError):
                    continue
        
        if not grouped_data:
            return None
            
        # Filter out very small values
        total = sum(grouped_data.values())
        if total <= 0:
            return None
            
        min_value = total * 0.01
        filtered_data = {k: v for k, v in grouped_data.items() if v >= min_value}
        
        if len(filtered_data) < 2:
            return None
            
        return filtered_data
    
    def identify_suitable_columns(self, headers, data):
        """Identify columns suitable for different types of charts."""
        categorical_cols = []
        numeric_cols = []
        
        for col in headers:
            if not col or col == "None":
                continue
                
            sample_values = []
            for row in data[:50]:  # Sample first 50 rows
                val = row.get(col)
                if val is not None:
                    sample_values.append(val)
            
            if not sample_values:
                continue
            
            # Check if numeric
            numeric_count = 0
            for val in sample_values:
                try:
                    if isinstance(val, (int, float)):
                        numeric_count += 1
                    else:
                        float(str(val).replace(',', '').replace('$', '').strip())
                        numeric_count += 1
                except:
                    pass
            
            if numeric_count / len(sample_values) > 0.7:  # 70% numeric
                numeric_cols.append(col)
            else:
                # Check uniqueness for categorical
                unique_vals = len(set(str(v) for v in sample_values))
                if 2 <= unique_vals <= 20:
                    categorical_cols.append(col)
        
        return categorical_cols, numeric_cols
    
    def process_xlsx_file(self, file_path):
        """Process a single XLSX file."""
        print(f"Processing: {file_path}")
        
        excel_data = self.read_xlsx_with_openpyxl(file_path)
        if not excel_data or not excel_data['data']:
            print(f"Skipping {file_path}: No data found")
            return
        
        headers = excel_data['headers']
        data = excel_data['data']
        file_name = Path(file_path).stem
        
        print(f"Data shape: {len(data)} rows, {len(headers)} columns")
        print(f"Columns: {headers}")
        
        # Identify suitable columns
        categorical_cols, numeric_cols = self.identify_suitable_columns(headers, data)
        print(f"Categorical columns: {categorical_cols}")
        print(f"Numeric columns: {numeric_cols}")
        
        charts_created = 0
        
        # Create pie charts for categorical columns
        for col in categorical_cols:
            chart_data = self.analyze_column_data(data, col)
            if chart_data:
                title = f"{file_name} - {col} Distribution"
                filename = self.output_dir / f"{file_name}_{col}_pie_chart.png"
                
                if self.create_pie_chart(chart_data, title, filename):
                    print(f"Created chart: {filename}")
                    charts_created += 1
        
        # Create pie charts for numeric columns grouped by categorical columns
        for num_col in numeric_cols:
            for cat_col in categorical_cols:
                grouped_data = self.process_numeric_by_category(data, num_col, cat_col)
                if grouped_data:
                    title = f"{file_name} - {num_col} by {cat_col}"
                    filename = self.output_dir / f"{file_name}_{num_col}_by_{cat_col}_pie_chart.png"
                    
                    if self.create_pie_chart(grouped_data, title, filename):
                        print(f"Created chart: {filename}")
                        charts_created += 1
        
        # Generate data summary
        summary_file = self.output_dir / f"{file_name}_data_summary.json"
        summary = {
            'file': str(file_path),
            'rows': len(data),
            'columns': headers,
            'categorical_columns': categorical_cols,
            'numeric_columns': numeric_cols,
            'charts_created': charts_created
        }
        
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=2)
        
        print(f"Created {charts_created} charts for {file_path}")
        print(f"Data summary saved to: {summary_file}")
        print("-" * 50)
    
    def process_all_xlsx_files(self, directory):
        """Process all XLSX files in a directory."""
        xlsx_files = list(Path(directory).glob("**/*.xlsx"))
        
        if not xlsx_files:
            print(f"No XLSX files found in {directory}")
            return
        
        print(f"Found {len(xlsx_files)} XLSX files")
        
        for xlsx_file in xlsx_files:
            self.process_xlsx_file(xlsx_file)


def main():
    """Main function."""
    parser = argparse.ArgumentParser(description="Generate pie charts from XLSX files")
    parser.add_argument(
        "--input-dir", 
        default="data", 
        help="Directory containing XLSX files (default: data)"
    )
    parser.add_argument(
        "--output-dir", 
        default="charts", 
        help="Output directory for generated charts (default: charts)"
    )
    parser.add_argument(
        "--file", 
        help="Process a specific XLSX file instead of a directory"
    )
    
    args = parser.parse_args()
    
    generator = SimplePieChartGenerator(args.output_dir)
    
    if args.file:
        if os.path.exists(args.file):
            generator.process_xlsx_file(args.file)
        else:
            print(f"File not found: {args.file}")
    else:
        if os.path.exists(args.input_dir):
            generator.process_all_xlsx_files(args.input_dir)
        else:
            print(f"Directory not found: {args.input_dir}")


if __name__ == "__main__":
    main()